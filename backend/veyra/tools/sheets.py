"""Spreadsheets — write real .xlsx files into the documents folder (openpyxl), so they
show in Files and open in Excel / LibreOffice.
"""

import time

from ..kernel import Tier, tool
from . import fsutil


def _slug(t):
    s = "".join(c.lower() if c.isalnum() else "-" for c in (t or "sheet"))
    while "--" in s:
        s = s.replace("--", "-")
    return s.strip("-") or "sheet"


@tool("sheets.make", Tier.CAUTION, "Create an Excel spreadsheet (.xlsx)",
      summarize=lambda a: f"Spreadsheet: {a.get('title', 'sheet')}",
      target=lambda a: f"documents/{a.get('title', 'sheet')}.xlsx")
def make_sheet(title="Sheet", columns=None, rows=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet")[:30]
    columns = columns or []
    rows = rows or []

    if columns:
        ws.append(columns)
        head_fill = PatternFill("solid", fgColor="D8693F")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
    for r in rows:
        ws.append(list(r))
    # auto width
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 50)

    name = f"{_slug(title)}-{int(time.time())}.xlsx"
    target = fsutil.resolve("documents", name)
    wb.save(str(target))
    return {"ok": True, "root": "documents", "path": name}
